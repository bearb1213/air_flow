"""
Script pour creer un fichier Excel
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

def create_excel(headers,data,filename):
    """
    Cree un fichier Excel avec des donnees d'exemple
    """
    # Creer un nouveau workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Donnees"
    
    # Ajouter des titres avec formatage
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajouter des donnees d'exemple
    
    
    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num)
            cell.value = value
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    
    # Sauvegarder le fichier
    wb.save(filename)
    print(f"✓ Fichier Excel cree avec succès: {filename}")

if __name__ == "__main__":
    header_traiteur = ["traiteur_code" , "client_code" , "nb_pax" , "date" , "lieu" , "status" ]
    data_traiteur = [
        # Client CLI001 - Fidèle (6 traiteurs)
        ['TRT001', 'CLI001', 50, '2024-01-15', 'Salle Eden Antsahavola', 'confirme'],
        ['TRT002', 'CLI001', 120, '2024-04-25', 'Salle Feon Razanakasina', 'confirme'],
        ['TRT003', 'CLI001', 105, '2024-09-25', 'Espace Fenoarivo', 'annule'],
        ['TRT004', 'CLI001', 85, '2024-03-25', 'Espace Anosibe', 'confirme'],
        ['TRT005', 'CLI001', 95, '2024-04-10', 'Salle Bezanozano', 'en attente'],
        ['TRT006', 'CLI001', 75, '2024-04-05', 'Domaine de Rova', 'annule'],
        
        # Client CLI002 - Fidèle (6 traiteurs)
        ['TRT007', 'CLI002', 75, '2024-01-20', 'Hotel Hilton', 'confirme'],
        ['TRT008', 'CLI002', 50, '2024-05-20', 'Villa paradis', 'en attente'],
        ['TRT009', 'CLI002', 80, '2024-08-05', 'Salle Espace Amitie', 'confirme'],
        ['TRT010', 'CLI002', 125, '2024-06-15', 'Chateau Belle Rive', 'confirme'],
        ['TRT011', 'CLI002', 130, '2024-06-08', 'Domaine Fleur', 'annule'],
        ['TRT012', 'CLI002', 110, '2024-11-15', 'Chateau Paradise', 'confirme'],
        
        # Client CLI003 - Fidèle (5 traiteurs)
        ['TRT013', 'CLI003', 100, '2024-01-25', 'Jardin tropical', 'confirme'],
        ['TRT014', 'CLI003', 85, '2024-05-01', 'Espace Cotiere', 'en attente'],
        ['TRT015', 'CLI003', 90, '2024-03-05', 'Salle Ambatondrazaka', 'en attente'],
        ['TRT016', 'CLI003', 75, '2024-10-01', 'Salle Feon Ambodifotatra', 'confirme'],
        ['TRT017', 'CLI003', 95, '2024-04-10', 'Salle Bezanozano', 'en attente'],
        
        # Client CLI004 - 3 traiteurs
        ['TRT018', 'CLI004', 60, '2024-02-01', 'Complexe sportif', 'en attente'],
        ['TRT019', 'CLI004', 130, '2024-08-10', 'Villa Anemone', 'confirme'],
        ['TRT020', 'CLI004', 80, '2024-11-20', 'Villa Rova', 'confirme'],
        
        # Client CLI005 - 3 traiteurs
        ['TRT021', 'CLI005', 110, '2024-05-05', 'Salle Anjou', 'confirme'],
        ['TRT022', 'CLI005', 140, '2024-10-05', 'Chateau Menalamba', 'confirme'],
        ['TRT023', 'CLI005', 75, '2024-05-22', 'Salle Ankadimbahoaka', 'annule'],
        
        # Client CLI006 - 2 traiteurs
        ['TRT024', 'CLI006', 45, '2024-02-10', 'Villa privee Antaninarenina', 'confirme'],
        ['TRT025', 'CLI006', 155, '2024-06-25', 'Complexe Hostellerie', 'confirme'],
        
        # Client CLI007 - 2 traiteurs
        ['TRT026', 'CLI007', 65, '2024-05-10', 'Domaine Tropical', 'confirme'],
        ['TRT027', 'CLI007', 200, '2024-04-01', 'Salle Crystal', 'confirme'],
        
        # Client CLI008 - 3 traiteurs
        ['TRT028', 'CLI008', 80, '2024-02-15', 'Parc Tsarasaotra', 'en attente'],
        ['TRT029', 'CLI008', 70, '2024-06-20', 'Salle Ankadimbahoaka', 'confirme'],
        ['TRT030', 'CLI008', 100, '2024-08-20', 'Chateau Rose', 'confirme'],
        
        # Client CLI009 - 3 traiteurs
        ['TRT031', 'CLI009', 150, '2024-02-25', 'Grand hotel Ivato', 'confirme'],
        ['TRT032', 'CLI009', 140, '2024-06-05', 'Salle Miarinarivo', 'confirme'],
        ['TRT033', 'CLI009', 125, '2024-10-15', 'Villa Cascade', 'en attente'],
        
        # Client CLI010 - 3 traiteurs
        ['TRT034', 'CLI010', 70, '2024-03-01', 'Lieux de reception', 'confirme'],
        ['TRT035', 'CLI010', 110, '2024-07-05', 'Parc de Loisirs', 'confirme'],
        ['TRT036', 'CLI010', 120, '2024-12-05', 'Domaine Nossi-Be', 'annule'],
        
        # Client CLI011 - 2 traiteurs
        ['TRT037', 'CLI011', 50, '2024-05-20', 'Villa paradis', 'en attente'],
        ['TRT038', 'CLI011', 95, '2024-10-20', 'Espace Amboasary', 'confirme'],
        
        # Client CLI012 - 2 traiteurs
        ['TRT039', 'CLI012', 70, '2024-06-20', 'Salle Ankadimbahoaka', 'confirme'],
        ['TRT040', 'CLI012', 185, '2024-07-10', 'Salle Ambohidahy', 'annule'],
        
        # Client CLI013 - 2 traiteurs
        ['TRT041', 'CLI013', 65, '2024-03-15', 'Villa au Lac', 'confirme'],
        ['TRT042', 'CLI013', 175, '2024-10-25', 'Salle Ravitavina', 'confirme'],
        
        # Client CLI014 - 2 traiteurs
        ['TRT043', 'CLI014', 130, '2024-03-20', 'Pavilion Mahamasina', 'en attente'],
        ['TRT044', 'CLI014', 60, '2024-10-10', 'Salle Behoririka', 'confirme'],
        
        # Client CLI015 - 2 traiteurs
        ['TRT045', 'CLI015', 85, '2024-03-25', 'Espace Anosibe', 'confirme'],
        ['TRT046', 'CLI015', 160, '2024-08-14', 'Salle Ramonatahy', 'annule'],
        
        # Client CLI016 - 2 traiteurs
        ['TRT047', 'CLI016', 75, '2024-04-05', 'Domaine de Rova', 'confirme'],
        ['TRT048', 'CLI016', 75, '2024-07-20', 'Salle Antananarivo', 'en attente'],
        
        # Client CLI017 - 2 traiteurs
        ['TRT049', 'CLI017', 100, '2024-05-25', 'Salle Raoul Follereau', 'confirme'],
        ['TRT050', 'CLI017', 150, '2024-11-05', 'Pavilion Fleur', 'en attente'],
        
        # Client CLI018 - 2 traiteurs
        ['TRT051', 'CLI018', 95, '2024-06-10', 'Jardin Botanique', 'en attente'],
        ['TRT052', 'CLI018', 165, '2024-07-25', 'Pavillon Voninkazo', 'confirme'],
        
        # Client CLI019 - 2 traiteurs
        ['TRT053', 'CLI019', 140, '2024-04-15', 'Chateau de Maison Rouge', 'confirme'],
        ['TRT054', 'CLI019', 70, '2024-09-19', 'Espace Lointain', 'annule'],
        
        # Client CLI020 - 2 traiteurs
        ['TRT055', 'CLI020', 55, '2024-04-20', 'Pavillon Floral', 'confirme'],
        ['TRT056', 'CLI020', 50, '2024-09-20', 'Domaine Tranquille', 'confirme'],
    ]
    filename_traiteur = "sql/traiteur.xlsx"

    create_excel(header_traiteur,data_traiteur,filename_traiteur)
    
    # =====================
    # MENUS TRAITEURS
    # =====================
    header_menu = ["traiteur_code", "type_plat", "plat_code", "label", "cout_preparation", "prix"]
    data_menu = [
        # TRT001 - 2 entrées, 3 plats principaux, 2 desserts
        ['TRT001', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT001', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT001', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT001', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT001', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT001', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT001', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        ['TRT001', 'Boisson', 'PLT020', 'Jus de fruits naturel', 2000.00, 4000.00],
        
        # TRT002 - 3 entrées, 4 plats principaux, 1 dessert
        ['TRT002', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT002', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT002', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT002', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT002', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT002', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT002', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT002', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        ['TRT002', 'Amuse-bouche', 'PLT025', 'Mini-brochettes aperitives x6', 4000.00, 7000.00],
        
        # TRT003 - 1 entrée, 5 plats principaux, 3 desserts
        ['TRT003', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT003', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT003', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT003', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT003', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT003', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT003', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT003', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        ['TRT003', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT004 - 2 entrées, 3 plats principaux, 2 desserts
        ['TRT004', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT004', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT004', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT004', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT004', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT004', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT004', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        
        # TRT005 - 0 entrée, 4 plats principaux, 1 dessert
        ['TRT005', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT005', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT005', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT005', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT005', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        ['TRT005', 'Cocktail', 'PLT027', 'Cocktail de bienvenue (verre)', 3000.00, 5000.00],
        
        # TRT006 - 3 entrées, 2 plats principaux, 0 dessert
        ['TRT006', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT006', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT006', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT006', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT006', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT006', 'Boisson', 'PLT022', 'Punch maison sans alcool', 3000.00, 5500.00],
        
        # TRT007 - 1 entrée, 5 plats principaux, 2 desserts
        ['TRT007', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT007', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT007', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT007', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT007', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT007', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT007', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT007', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT008 - 2 entrées, 3 plats principaux, 3 desserts
        ['TRT008', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT008', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT008', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT008', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT008', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT008', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT008', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        ['TRT008', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT009 - 0 entrée, 4 plats principaux, 2 desserts
        ['TRT009', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT009', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT009', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT009', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT009', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        ['TRT009', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT010 - 3 entrées, 2 plats principaux, 1 dessert
        ['TRT010', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT010', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT010', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT010', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT010', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT010', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT010', 'Amuse-bouche', 'PLT026', 'Verrines legumes x4', 3500.00, 6000.00],
        
        # TRT011 - 1 entrée, 3 plats principaux, 2 desserts
        ['TRT011', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT011', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT011', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT011', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT011', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT011', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        
        # TRT012 - 2 entrées, 5 plats principaux, 1 dessert
        ['TRT012', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT012', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT012', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT012', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT012', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT012', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT012', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT012', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT013 - 0 entrée, 3 plats principaux, 3 desserts
        ['TRT013', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT013', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT013', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT013', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT013', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        ['TRT013', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT014 - 3 entrées, 4 plats principaux, 0 dessert
        ['TRT014', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT014', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT014', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT014', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT014', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT014', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT014', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT014', 'Boisson', 'PLT023', 'Cafe / The', 1000.00, 2500.00],
        
        # TRT015 - 2 entrées, 2 plats principaux, 2 desserts
        ['TRT015', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT015', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT015', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT015', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT015', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT015', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT016 - 1 entrée, 5 plats principaux, 1 dessert
        ['TRT016', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT016', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT016', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT016', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT016', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT016', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT016', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        
        # TRT017 - 3 entrées, 3 plats principaux, 2 desserts
        ['TRT017', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT017', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT017', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT017', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT017', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT017', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT017', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT017', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT018 - 0 entrée, 2 plats principaux, 3 desserts
        ['TRT018', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT018', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT018', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT018', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        ['TRT018', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT019 - 2 entrées, 4 plats principaux, 2 desserts
        ['TRT019', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT019', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT019', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT019', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT019', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT019', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT019', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT019', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        
        # TRT020 - 1 entrée, 3 plats principaux, 1 dessert
        ['TRT020', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT020', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT020', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT020', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT020', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        
        # TRT021 - 2 entrées, 3 plats principaux, 2 desserts
        ['TRT021', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT021', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT021', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT021', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT021', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT021', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT021', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT022 - 1 entrée, 4 plats principaux, 2 desserts
        ['TRT022', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT022', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT022', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT022', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT022', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT022', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        ['TRT022', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT023 - 3 entrées, 2 plats principaux, 1 dessert
        ['TRT023', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT023', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT023', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT023', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT023', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT023', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        
        # TRT024 - 0 entrée, 3 plats principaux, 3 desserts
        ['TRT024', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT024', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT024', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT024', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT024', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        ['TRT024', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT025 - 2 entrées, 4 plats principaux, 1 dessert
        ['TRT025', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT025', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT025', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT025', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT025', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT025', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT025', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT026 - 1 entrée, 5 plats principaux, 2 desserts
        ['TRT026', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT026', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT026', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT026', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT026', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT026', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT026', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT026', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        
        # TRT027 - 3 entrées, 3 plats principaux, 1 dessert
        ['TRT027', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT027', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT027', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT027', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT027', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT027', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT027', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        
        # TRT028 - 2 entrées, 2 plats principaux, 3 desserts
        ['TRT028', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT028', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT028', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT028', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT028', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT028', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        ['TRT028', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT029 - 1 entrée, 3 plats principaux, 2 desserts
        ['TRT029', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT029', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT029', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT029', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT029', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT029', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT030 - 3 entrées, 2 plats principaux, 0 dessert
        ['TRT030', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT030', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT030', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT030', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT030', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        
        # TRT031 - 0 entrée, 4 plats principaux, 2 desserts
        ['TRT031', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT031', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT031', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT031', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT031', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        ['TRT031', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT032 - 2 entrées, 3 plats principaux, 2 desserts
        ['TRT032', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT032', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT032', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT032', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT032', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT032', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT032', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        
        # TRT033 - 1 entrée, 4 plats principaux, 1 dessert
        ['TRT033', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT033', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT033', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT033', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT033', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT033', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT034 - 3 entrées, 3 plats principaux, 2 desserts
        ['TRT034', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT034', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT034', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT034', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT034', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT034', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT034', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT034', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT035 - 2 entrées, 2 plats principaux, 2 desserts
        ['TRT035', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT035', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT035', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT035', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT035', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        ['TRT035', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT036 - 1 entrée, 5 plats principaux, 1 dessert
        ['TRT036', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT036', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT036', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT036', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT036', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT036', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT036', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        
        # TRT037 - 2 entrées, 3 plats principaux, 1 dessert
        ['TRT037', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT037', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT037', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT037', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT037', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT037', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        
        # TRT038 - 0 entrée, 3 plats principaux, 3 desserts
        ['TRT038', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT038', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT038', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT038', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT038', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT038', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        
        # TRT039 - 3 entrées, 2 plats principaux, 1 dessert
        ['TRT039', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT039', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT039', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT039', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT039', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT039', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT040 - 1 entrée, 4 plats principaux, 2 desserts
        ['TRT040', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT040', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT040', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT040', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT040', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT040', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT040', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        
        # TRT041 - 2 entrées, 3 plats principaux, 2 desserts
        ['TRT041', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT041', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT041', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT041', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT041', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT041', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT041', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT042 - 3 entrées, 2 plats principaux, 2 desserts
        ['TRT042', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT042', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT042', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT042', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT042', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT042', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        ['TRT042', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT043 - 2 entrées, 2 plats principaux, 3 desserts
        ['TRT043', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT043', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT043', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT043', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT043', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT043', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        ['TRT043', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT044 - 1 entrée, 3 plats principaux, 1 dessert
        ['TRT044', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT044', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT044', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT044', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT044', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        
        # TRT045 - 2 entrées, 4 plats principaux, 1 dessert
        ['TRT045', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT045', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT045', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT045', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT045', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT045', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT045', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        
        # TRT046 - 3 entrées, 3 plats principaux, 1 dessert
        ['TRT046', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT046', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT046', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT046', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT046', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT046', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT046', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        
        # TRT047 - 1 entrée, 4 plats principaux, 2 desserts
        ['TRT047', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT047', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT047', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT047', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT047', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT047', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        ['TRT047', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT048 - 2 entrées, 3 plats principaux, 2 desserts
        ['TRT048', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT048', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT048', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT048', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT048', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT048', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT048', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT049 - 0 entrée, 4 plats principaux, 2 desserts
        ['TRT049', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT049', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT049', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT049', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT049', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT049', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        
        # TRT050 - 3 entrées, 2 plats principaux, 2 desserts
        ['TRT050', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT050', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT050', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT050', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT050', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT050', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        ['TRT050', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT051 - 2 entrées, 3 plats principaux, 1 dessert
        ['TRT051', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT051', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT051', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT051', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT051', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT051', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        
        # TRT052 - 1 entrée, 5 plats principaux, 2 desserts
        ['TRT052', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT052', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT052', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT052', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT052', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT052', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT052', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT052', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
        
        # TRT053 - 2 entrées, 2 plats principaux, 2 desserts
        ['TRT053', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT053', 'Entree', 'PLT005', 'Soupe de tomate basilic', 3500.00, 7000.00],
        ['TRT053', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT053', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT053', 'Dessert', 'PLT016', 'Mousse au chocolat', 5000.00, 9000.00],
        ['TRT053', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        
        # TRT054 - 3 entrées, 3 plats principaux, 1 dessert
        ['TRT054', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT054', 'Entree', 'PLT003', 'Carpaccio de thon', 7000.00, 13000.00],
        ['TRT054', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT054', 'Plat principal', 'PLT007', 'Poulet grille sauce coco', 10000.00, 18000.00],
        ['TRT054', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT054', 'Plat principal', 'PLT012', 'Lasagnes bolognaise', 10000.00, 17000.00],
        ['TRT054', 'Dessert', 'PLT019', 'Tarte aux fraises', 7000.00, 12000.00],
        
        # TRT055 - 1 entrée, 3 plats principaux, 2 desserts
        ['TRT055', 'Entree', 'PLT004', 'Salade avocat crevettes', 8000.00, 14000.00],
        ['TRT055', 'Plat principal', 'PLT006', 'Romazava traditionnel', 8000.00, 15000.00],
        ['TRT055', 'Plat principal', 'PLT008', 'Riz cantonnais', 5000.00, 9000.00],
        ['TRT055', 'Plat principal', 'PLT010', 'Poisson sauce gingembre', 11000.00, 20000.00],
        ['TRT055', 'Dessert', 'PLT014', 'Gâteau vanille maison', 6000.00, 10000.00],
        ['TRT055', 'Dessert', 'PLT017', 'Panna cotta fruits rouges', 5500.00, 9500.00],
        
        # TRT056 - 2 entrées, 3 plats principaux, 2 desserts
        ['TRT056', 'Entree', 'PLT001', 'Salade de crudites', 3000.00, 6000.00],
        ['TRT056', 'Entree', 'PLT002', 'Veloute de legumes', 4000.00, 7500.00],
        ['TRT056', 'Plat principal', 'PLT009', 'Zebu braise sauce poivre vert', 14000.00, 25000.00],
        ['TRT056', 'Plat principal', 'PLT011', 'Brochettes de porc marine', 9000.00, 16000.00],
        ['TRT056', 'Plat principal', 'PLT013', 'Poulet rôti pommes de terre', 12000.00, 21000.00],
        ['TRT056', 'Dessert', 'PLT015', 'Salade de fruits frais', 3500.00, 6500.00],
        ['TRT056', 'Dessert', 'PLT018', 'Creme brûlee', 4500.00, 8500.00],
    ]
    filename_menu = "sql/menu_traiteur.xlsx"

    create_excel(header_menu,data_menu,filename_menu)
    print("✓ Tous les fichiers Excel ont été créés avec succès!")

    
   